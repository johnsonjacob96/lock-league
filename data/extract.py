"""Extract Lock League data from XLSX into structured JSON.

Outputs to /tmp/lock-league/data.json with shape:
{
  "seasons": {
    "2023": { "members": [...], "weeklyWinners": {1: "Mason", ...}, "leagueAllTime": "..." },
    "2024": {...},
    "2025": {...}
  },
  "members": {
    "Jacob": {
      "2025": { "seasonRecord": "47-38", "lifetimeRecord": "119-124-2",
                "byType": { "Favorite": {"record": "8-5-0", "picks": {1: {text, result}, ...}},
                            "Dog": {...}, "Over": {...}, "Under": {...}, "Super Lock": {...} } },
      "2024": {...}, "2023": {...}
    },
    ...
  },
  "rules": [{"article": "Article 1: ...", "amended": "Week 1, 2023"}, ...],
  "parlays": { "2023": [...], "2024": [...], "2025": [...] }
}
"""
import json
import openpyxl
from collections import OrderedDict

WIN_COLORS = {"FF93C47D", "FFB6D7A8"}
LOSS_COLORS = {"FFE06666"}
PUSH_COLORS = {"FFFFD966"}

def classify(fg):
    if fg in WIN_COLORS:
        return "W"
    if fg in LOSS_COLORS:
        return "L"
    if fg in PUSH_COLORS:
        return "P"
    return None

def cell_fg(cell):
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    try:
        return fill.fgColor.rgb
    except Exception:
        return None

BET_TYPE_ROW_OFFSET = {
    "Favorite": 1,
    "Dog": 2,
    "Over": 3,
    "Under": 4,
}
SUPER_LOCK_NAMES = {"Super 🔒", "Super 🔐", "Super Lock"}

def parse_bets_sheet(ws, season_label):
    """Return list of member dicts for a season Bets sheet."""
    members = []
    max_row = ws.max_row
    max_col = ws.max_column

    # Find member header rows: cell (r, 2) == "Bet Type"
    header_rows = []
    for r in range(1, max_row + 1):
        if ws.cell(row=r, column=2).value == "Bet Type":
            header_rows.append(r)

    # For each header row: extract member block (rows r..r+5)
    for hr in header_rows:
        name = ws.cell(row=hr, column=1).value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if name in ("|\\\\\\\\\\\\\\\\\\\\", "a"):
            continue

        # Identify week headers from header row, cols 5+
        week_labels = OrderedDict()  # col -> week number
        for c in range(5, max_col + 1):
            v = ws.cell(row=hr, column=c).value
            if isinstance(v, str) and v.upper().startswith("WEEK"):
                # parse number
                tokens = v.replace("WEEK", "").strip().split()
                if tokens and tokens[0].isdigit():
                    week_labels[c] = int(tokens[0])

        # Sidebar: col 1 of rows hr+1..hr+5
        sidebar = {}
        for off in range(1, 6):
            r = hr + off
            v = ws.cell(row=r, column=1).value
            if v is not None:
                sidebar[off] = v

        # In the 2025 sheet observed:
        # off=2 → "2025 Season" label, off=3 → season record, off=4 → "Lifetime", off=5 → lifetime record
        # But values can shift slightly between members; scan generically
        season_record = None
        lifetime_record = None
        for off, v in sidebar.items():
            sv = str(v).strip()
            if "Season" in sv:
                # next slot is season record
                nxt = sidebar.get(off + 1)
                if nxt:
                    season_record = str(nxt).strip()
            elif sv.lower() == "lifetime":
                nxt = sidebar.get(off + 1)
                if nxt:
                    lifetime_record = str(nxt).strip()
            elif "-" in sv and any(ch.isdigit() for ch in sv) and season_record is None and lifetime_record is None:
                # fallback: bare record string
                pass

        # Bet type rows: scan col 2 in rows hr+1..hr+5
        by_type = OrderedDict()
        for off in range(1, 6):
            r = hr + off
            btype = ws.cell(row=r, column=2).value
            if not btype:
                continue
            btype = str(btype).strip()
            # Normalize Super Lock
            if btype in SUPER_LOCK_NAMES or btype.startswith("Super"):
                btype = "Super Lock"
            record = ws.cell(row=r, column=3).value
            picks = OrderedDict()
            for c, wk in week_labels.items():
                cell = ws.cell(row=r, column=c)
                text = cell.value
                if isinstance(text, str):
                    text = text.strip()
                fg = cell_fg(cell)
                result = classify(fg)
                if text or result:
                    picks[wk] = {
                        "text": text if text else None,
                        "result": result,
                        "fg": fg,
                    }
            by_type[btype] = {
                "record": str(record).strip() if record else None,
                "picks": picks,
            }

        members.append({
            "name": name,
            "season": season_label,
            "seasonRecord": season_record,
            "lifetimeRecord": lifetime_record,
            "byType": by_type,
        })

    return members

def parse_weekly_winners(ws):
    """Return {season: {week: winner}}."""
    out = {}
    max_row = ws.max_row
    max_col = ws.max_column
    # Row 1 has season labels at columns where season blocks start; row 2 has "Week" / "Winner" pairs
    season_cols = {}  # season_label -> col of "Week" column
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and "Season" in v:
            season_cols[v.strip()] = c
    # For each season, columns are (week_col, winner_col) = (c, c+1)
    for season, wc in season_cols.items():
        winners = OrderedDict()
        for r in range(3, max_row + 1):
            wk_cell = ws.cell(row=r, column=wc).value
            win_cell = ws.cell(row=r, column=wc + 1).value
            if not wk_cell:
                continue
            wk_str = str(wk_cell).strip()
            if wk_str.lower().startswith("week"):
                tok = wk_str.split()
                if len(tok) >= 2 and tok[1].isdigit():
                    winners[int(tok[1])] = str(win_cell).strip() if win_cell else None
        # season label like "2023 Season" -> year key
        year = season.split()[0]
        out[year] = winners
    return out

def parse_rules(ws):
    rules = []
    for r in range(2, ws.max_row + 1):
        article = ws.cell(row=r, column=1).value
        amended = ws.cell(row=r, column=2).value
        if article and isinstance(article, str) and article.strip():
            rules.append({"article": article.strip(), "amended": str(amended).strip() if amended else None})
    return rules

def main():
    wb = openpyxl.load_workbook("/tmp/lock-league/lock-league.xlsx", data_only=True)
    data = {"seasons": {}, "members": {}, "rules": [], "parlays": {}}

    # Rules
    if "Rules" in wb.sheetnames:
        data["rules"] = parse_rules(wb["Rules"])

    # Weekly winners
    weekly_winners = {}
    if "Weekly Winners" in wb.sheetnames:
        weekly_winners = parse_weekly_winners(wb["Weekly Winners"])

    # Bet sheets per year
    for year, sheet_name in [("2023", "2023 Bets"), ("2024", "2024 Bets"), ("2025", "2025 Bets")]:
        if sheet_name not in wb.sheetnames:
            continue
        members_in_year = parse_bets_sheet(wb[sheet_name], year)
        # Collect distinct fg values observed for diagnostic
        fg_seen = {}
        for m in members_in_year:
            for bt, info in m["byType"].items():
                for wk, pick in info["picks"].items():
                    fg = pick.get("fg")
                    if fg:
                        fg_seen[fg] = fg_seen.get(fg, 0) + 1
        data["seasons"][year] = {
            "members": [m["name"] for m in members_in_year],
            "weeklyWinners": weekly_winners.get(year, {}),
            "fgCounts": fg_seen,
        }
        for m in members_in_year:
            data["members"].setdefault(m["name"], {})[year] = {
                "seasonRecord": m["seasonRecord"],
                "lifetimeRecord": m["lifetimeRecord"],
                "byType": m["byType"],
            }

    # Strip fg fields from final output (kept only for diagnostic)
    for member, years in data["members"].items():
        for year, info in years.items():
            for bt, btinfo in info["byType"].items():
                for wk, pick in btinfo["picks"].items():
                    pick.pop("fg", None)

    out_path = "/tmp/lock-league/data.json"
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Wrote {out_path}")
    # Diagnostic summary
    print("Members:", sorted(data["members"].keys()))
    for y in sorted(data["seasons"].keys()):
        s = data["seasons"][y]
        print(f"  {y}: {len(s['members'])} members, {len(s['weeklyWinners'])} weekly winners, fgs={s['fgCounts']}")
    print(f"Rules: {len(data['rules'])}")

if __name__ == "__main__":
    main()
